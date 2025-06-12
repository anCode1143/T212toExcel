import os
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from AccountData import get_cash_info, get_open_positions, get_pies

CACHE_DIR = "cache"

def load_cached(name, fallback_func):
    path = os.path.join(CACHE_DIR, f"{name}.json")
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return fallback_func()

def make_xslx():
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Summary"
    
    # Create an empty "Advanced Account Info" sheet
    wb.create_sheet("Advanced Account Info")

    # Styles
    table_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    dark_grey = PatternFill(start_color="9C9C9C", end_color="9C9C9C", fill_type="solid")
    grey = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
    red = PatternFill(start_color="e8baba", end_color="e8baba", fill_type="solid")
    green = PatternFill(start_color="c3e8cb", end_color="c3e8cb", fill_type="solid")

    def cash_info_table():
        cash_info = load_cached("cash_info", get_cash_info)
        
        # Calculate total fees from CSV file
        total_fees = 0
        csv_path = "trading212_history.csv"
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    # Deposit fees
                    deposit_fee = row.get("Deposit fee", "0")
                    if deposit_fee and deposit_fee != "0":
                        try:
                            total_fees += abs(float(deposit_fee))
                        except ValueError:
                            pass
                    
                    # Currency conversion fees
                    conv_fee = row.get("Currency conversion fee", "0")
                    if conv_fee and conv_fee != "0":
                        try:
                            total_fees += abs(float(conv_fee))
                        except ValueError:
                            pass
                    
                    # Charge amounts (trading fees)
                    charge_amount = row.get("Charge amount", "0")
                    if charge_amount and charge_amount != "0":
                        try:
                            total_fees += abs(float(charge_amount))
                        except ValueError:
                            pass
                    
                    # Stamp duty reserve tax
                    stamp_duty = row.get("Stamp duty reserve tax", "0")
                    if stamp_duty and stamp_duty != "0":
                        try:
                            total_fees += abs(float(stamp_duty))
                        except ValueError:
                            pass
        
        # Add calculated fees to cash_info
        cash_info["fees"] = round(total_fees, 2)
        
        title_range = "B2:D2"
        ws.merge_cells(title_range)
        title_cell = ws['B2']
        title_cell.value = "Cash Info"
        title_cell.font = Font(bold=True)
        title_cell.fill = dark_grey
        for row in ws[title_range]:
            for cell in row:
                cell.border = title_border

        label_map = {
            "total": "Total Account Value",
            "free": "Cash",
            "invested": "Currently Invested",
            "blocked": "Blocked Amount",
            "pieCash": "Cash in Pies",
            "result": "Realised Return",
            "ppl": "Open Position P/L",
            "fees": "Total Fees Paid"
        }

        row = 3
        first_row = row
        last_row = row + len(label_map) - 1

        for key, label in label_map.items():
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            label_cell = ws.cell(row=row, column=2, value=label)
            value = cash_info.get(key) or 0
            value_cell = ws.cell(row=row, column=4, value=value)
            label_cell.fill = grey
            value_cell.fill = grey
            if key in ["result", "ppl"] and isinstance(value, (int, float)):
                if value > 0:
                    label_cell.fill = green
                    value_cell.fill = green
                elif value < 0:
                    label_cell.fill = red
                    value_cell.fill = red
            elif key == "fees" and isinstance(value, (int, float)) and value > 0:
                # Display fees in red since they represent costs
                label_cell.fill = red
                value_cell.fill = red
            for col in range(2, 5):
                ws.cell(row=row, column=col).border = table_border
            row += 1
        
        # Apply full table border
        thin_side = Side(style='thin')
        for r in range(2, last_row + 1):
            for c in range(2, 5):
                cell = ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == 2 else border.left,
                    right=thin_side if c == 4 else border.right,
                    top=thin_side if r == 2 else border.top,
                    bottom=thin_side if r == last_row else border.bottom,
                )
                cell.border = new_border

    def open_positions_table():
        positions = load_cached("open_positions", get_open_positions)
        positions = sorted(positions, key=lambda x: x.get("ppl", 0), reverse=True)
        start_col = 6
        start_row = 2
        
        # Title
        title_range = "F2:K2"
        ws.merge_cells(title_range)
        title_cell = ws['F2']
        title_cell.value = "Open Positions"
        title_cell.font = Font(bold=True)
        title_cell.fill = dark_grey
        for row in ws[title_range]:
            for cell in row:
                cell.border = title_border
        
        # Headers
        headers = ["Ticker", "Quantity", "Avg. Price", "Current Price", "P/L", "FX P/L"]
        header_row = start_row + 1
        for col_offset, header in enumerate(headers):
            col = start_col + col_offset
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = grey
            cell.border = table_border
            cell.font = Font(bold=True)
        
        # Data rows
        row = header_row + 1
        for pos in positions:
            ticker = pos.get("ticker", "N/A").split("_")[0]
            quantity = round(pos.get("quantity") or 0.0, 2)
            avg_price = round(pos.get("averagePrice") or 0.0, 2)
            current_price = round(pos.get("currentPrice") or 0.0, 2)
            ppl = round(pos.get("ppl") or 0.0, 2)
            fx_ppl = round(pos.get("fxPpl") or 0.0, 2)
            values = [ticker, quantity, avg_price, current_price, ppl, fx_ppl]
            row_fill = green if ppl > 0 else red if ppl < 0 else grey
            
            for col_offset, val in enumerate(values):
                col = start_col + col_offset
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = table_border
                if col_offset == 0:
                    cell.fill = grey
                elif col_offset == 5:
                    if fx_ppl > 0:
                        cell.fill = green
                    elif fx_ppl < 0:
                        cell.fill = red
                    else:
                        cell.fill = grey
                else:
                    cell.fill = row_fill
            row += 1
        
        # Set column widths
        for col_letter in ['F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col_letter].width = 15
        
        # Apply full table border
        thin_side = Side(style='thin')
        first_row = start_row
        last_data_row = row - 1
        first_col = start_col
        last_col = start_col + len(headers) - 1
        for r in range(first_row, last_data_row + 1):
            for c in range(first_col, last_col + 1):
                cell = ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == first_col else border.left,
                    right=thin_side if c == last_col else border.right,
                    top=thin_side if r == first_row else border.top,
                    bottom=thin_side if r == last_data_row else border.bottom,
                )
                cell.border = new_border

    def historical_transactions():
        # Read transactions from CSV file instead of cache
        transactions_info = []
        csv_path = "trading212_history.csv"
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    action = row.get("Action", "")
                    # Only include deposit and withdrawal transactions
                    if action.lower() in ["deposit", "withdrawal", "withdraw"]:
                        # Parse the transaction data
                        time_str = row.get("Time", "")
                        date = time_str.split(" ")[0] if " " in time_str else time_str
                        total = row.get("Total", "0")
                        
                        # Convert total to float, handle empty or non-numeric values
                        try:
                            amount = float(total) if total else 0
                        except ValueError:
                            amount = 0
                        
                        transactions_info.append({
                            "dateTime": time_str,
                            "type": action,
                            "amount": amount
                        })
        
        start_col = 2
        start_row = 11
        
        # Title
        title_range = "B11:D11"
        ws.merge_cells(title_range)
        title_cell = ws['B11']
        title_cell.value = "Total Transactions"
        title_cell.font = Font(bold=True)
        title_cell.fill = dark_grey
        for row in ws[title_range]:
            for cell in row:
                cell.border = title_border
        
        # Headers
        subheader_row = start_row + 1
        subheaders = ["Date", "Transaction Type", "Amount"]
        for col_offset, subheader in enumerate(subheaders):
            col = start_col + col_offset
            cell = ws.cell(row=subheader_row, column=col, value=subheader)
            cell.font = Font(bold=True)
            cell.fill = grey
            cell.border = table_border
        
        # Data rows
        row = subheader_row + 1
        for tx in transactions_info:
            tx_type = tx.get("type", "N/A")
            if str(tx_type).lower() == "fee":
                continue
            date_time = tx.get("dateTime", "")
            # Extract just the date part, removing time for readability
            if "T" in date_time:
                date = date_time.split("T")[0]
            elif " " in date_time:
                date = date_time.split(" ")[0]
            else:
                date = date_time
            amount = tx.get("amount", 0)
            tx_type_lower = str(tx_type).lower()
            
            if tx_type_lower == "deposit":
                right_fill = green
            elif tx_type_lower == "withdraw" or tx_type_lower == "withdrawal":
                right_fill = red
            else:
                right_fill = grey
            
            values = [date, tx_type, amount]
            for col_offset, val in enumerate(values):
                col = start_col + col_offset
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = table_border
                if col_offset == 1 or col_offset == 2:
                    cell.fill = right_fill
                else:
                    cell.fill = grey
            row += 1
        
        # Set column widths
        for col_letter in ['B', 'C', 'D']:
            ws.column_dimensions[col_letter].width = 15
        
        # Apply full table border
        thin_side = Side(style='thin')
        first_row = start_row
        last_data_row = row - 1
        first_col = start_col
        last_col = start_col + len(subheaders) - 1
        for r in range(first_row, last_data_row + 1):
            for c in range(first_col, last_col + 1):
                cell = ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == first_col else border.left,
                    right=thin_side if c == last_col else border.right,
                    top=thin_side if r == first_row else border.top,
                    bottom=thin_side if r == last_data_row else border.bottom,
                )
                cell.border = new_border

    def pies_tables():
        pies_info = load_cached("pies_info", lambda: get_pies(include_detailed=True))
        start_col = 13  # Column M (beside open positions)
        start_row = 2
        
        for pie in pies_info:
            pie_id = pie.get("id", "")
            name = pie.get("detailed", {}).get("settings", {}).get("name") or pie.get("name", "N/A")
            detailed = pie.get("detailed", {})
            instruments = detailed.get("instruments", [])
            
            # Skip pies with no assets or value
            if not instruments:
                continue
            total_value = sum(inst.get("result", {}).get("priceAvgValue", 0) for inst in instruments)
            if total_value == 0:
                continue
            
            # Sort instruments by performance (greatest gains to greatest losses)
            instruments = sorted(instruments, key=lambda x: x.get("result", {}).get("priceAvgResultCoef", 0), reverse=True)
            
            # Get pie summary data
            pie_result = pie.get("result", {})
            total_invested = round(pie_result.get("priceAvgInvestedValue", 0), 2)
            pie_pl = round(pie_result.get("priceAvgResult", 0), 2)
            pie_pl_percent = round(pie_result.get("priceAvgResultCoef", 0) * 100, 2)
            
            # Title
            title_range = ws.cell(row=start_row, column=start_col).coordinate + ":" + ws.cell(row=start_row, column=start_col+4).coordinate
            ws.merge_cells(title_range)
            title_cell = ws.cell(row=start_row, column=start_col)
            title_cell.value = f"Pie: {name} (ID: {pie_id})"
            title_cell.font = Font(bold=True)
            title_cell.fill = dark_grey
            for col in range(start_col, start_col+5):
                cell = ws.cell(row=start_row, column=col)
                cell.border = title_border
            
            # Holdings header
            holdings_row = start_row + 1
            subheaders = ["Ticker", "Weight %", "Performance %", "Quantity", "Value"]
            for col_offset, subheader in enumerate(subheaders):
                col = start_col + col_offset
                cell = ws.cell(row=holdings_row, column=col, value=subheader)
                cell.font = Font(bold=True)
                cell.fill = grey
                cell.border = table_border
            
            # Holdings data
            row = holdings_row + 1
            for inst in instruments:
                ticker = inst.get("ticker", "").split("_")[0]  # Clean ticker
                weight = round(inst.get("currentShare", 0) * 100, 2)
                perf = round(inst.get("result", {}).get("priceAvgResultCoef", 0) * 100, 2)
                qty = round(inst.get("ownedQuantity", 0), 4)
                value = round(inst.get("result", {}).get("priceAvgValue", 0), 2)
                values = [ticker, weight, perf, qty, value]
                row_fill = green if perf > 0 else red if perf < 0 else grey
                
                for col_offset, val in enumerate(values):
                    col = start_col + col_offset
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = table_border
                    if col_offset == 0:  # Ticker column
                        cell.fill = grey
                    else:
                        cell.fill = row_fill
                row += 1
            
            # Total value row
            total_row = row
            ws.cell(row=total_row, column=start_col, value="Total Pie Value:").font = Font(bold=True)
            ws.cell(row=total_row, column=start_col+4, value=round(total_value, 2)).font = Font(bold=True)
            for col in range(start_col, start_col+5):
                cell = ws.cell(row=total_row, column=col)
                cell.border = table_border
                cell.fill = grey
            
            # Pie Summary Section
            summary_row = total_row + 1
            summary_labels = ["Initial Investment:", "Pie P/L:", "P/L %:"]
            summary_values = [total_invested, pie_pl, pie_pl_percent]
            
            for i, (label, value) in enumerate(zip(summary_labels, summary_values)):
                # Create label in first column and value in last column
                ws.cell(row=summary_row + i, column=start_col, value=label).font = Font(bold=True)
                ws.cell(row=summary_row + i, column=start_col + 4, value=value).font = Font(bold=True)
                
                # Determine value fill color for P/L rows
                if i > 0:  # P/L and P/L % rows
                    if value > 0:
                        value_fill = green
                    elif value < 0:
                        value_fill = red
                    else:
                        value_fill = grey
                else:
                    value_fill = grey
                
                # Apply styling and borders to all cells in the row
                for col in range(start_col, start_col+5):
                    cell = ws.cell(row=summary_row + i, column=col)
                    cell.border = table_border
                    if col == start_col + 4:  # Value column
                        cell.fill = value_fill
                    else:  # Label and middle columns
                        cell.fill = grey
            
            last_summary_row = summary_row + len(summary_labels) - 1
            
            # Apply full table border
            thin_side = Side(style='thin')
            first_row = start_row
            last_data_row = last_summary_row
            first_col = start_col
            last_col = start_col + 4
            for r in range(first_row, last_data_row + 1):
                for c in range(first_col, last_col + 1):
                    cell = ws.cell(row=r, column=c)
                    border = cell.border
                    new_border = Border(
                        left=thin_side if c == first_col else border.left,
                        right=thin_side if c == last_col else border.right,
                        top=thin_side if r == first_row else border.top,
                        bottom=thin_side if r == last_data_row else border.bottom,
                    )
                    cell.border = new_border
            
            # Set column widths
            for col_idx in range(start_col, start_col+5):
                col_letter = chr(64 + col_idx)
                ws.column_dimensions[col_letter].width = 15
            
            start_row = last_summary_row + 3  # Add space between pies

    # Execute all table functions
    cash_info_table()
    open_positions_table()
    historical_transactions()
    pies_tables()

    # Save the workbook
    wb.save("AccountAnalysis.xlsx")

if __name__ == "__main__":
    make_xslx()