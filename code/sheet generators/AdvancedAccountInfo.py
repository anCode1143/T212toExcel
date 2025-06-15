import os
import csv
import yfinance as yf
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Rectangle
import io
from openpyxl.drawing.image import Image
from collections import defaultdict

class AdvancedAccountInfo:
    def __init__(self, wb, ws, styles, extract_date_func, apply_border_func):
        self.wb = wb
        self.ws = ws
        self.styles = styles
        self.extract_date = extract_date_func
        self.apply_table_border = apply_border_func
        
    def order_history(self):
        transactions_info = []
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        # Create ticker to ISIN and currency mapping from trading history CSV
        ticker_to_isin = {}
        ticker_to_currency = {}
        
        def is_uk_security_in_pence(isin, ticker, trading_currency=None):
            """Check if a security is quoted in pence, requiring conversion."""
            # First check if we have trading currency info from T212 history
            if trading_currency == "GBX":
                return True
            
            # Then check if it's a UK security
            if not isin or not isin.startswith("GB"):
                return False
            
            try:
                import yfinance as yf
                # Convert T212 ticker to Yahoo Finance ticker format
                # Remove the suffix and add .L for London Stock Exchange
                base_ticker = ticker.split("_")[0]
                
                # Common T212 to Yahoo Finance ticker mappings for UK securities
                ticker_mappings = {
                    "PSN": "PSN.L",     # Persimmon
                    "SVS": "SVS.L",     # Savills  
                    "TW": "TW.L",       # Taylor Wimpey
                    "BLND": "BLND.L",   # British Land
                    "COPAP": "COPA.L",  # WisdomTree Copper (might be different)
                    "OD7Z": "ODGD.L",   # WisdomTree Industrial Metals (might be different)
                    "SUGA": "SUGA.L",   # WisdomTree Sugar (might be different)
                    "AIGAP": "AIGA.L",  # WisdomTree Agriculture (might be different)
                }
                
                # Use mapping if available, otherwise try adding .L
                if base_ticker in ticker_mappings:
                    yahoo_ticker = ticker_mappings[base_ticker]
                else:
                    # Remove trailing 'l' or 'L' if present and add .L
                    if base_ticker.endswith(('l', 'L')):
                        base_ticker = base_ticker[:-1]
                    yahoo_ticker = f"{base_ticker}.L"
                
                # Suppress yfinance and HTTP library output and errors
                import warnings, logging, requests
                warnings.filterwarnings("ignore", message="Unverified HTTPS request")
                requests.packages.urllib3.disable_warnings()
                logging.getLogger("yfinance").setLevel(logging.CRITICAL)
                logging.getLogger("urllib3").setLevel(logging.CRITICAL)
                logging.getLogger("requests").setLevel(logging.CRITICAL)
                stock = yf.Ticker(yahoo_ticker)
                info = stock.info
                currency = info.get("currency", "")
                
                # If currency is GBp (pence), we need to convert to pounds
                return currency == "GBp"
            except Exception as e:
                print(f"Warning: Could not determine currency for {ticker} ({yahoo_ticker if 'yahoo_ticker' in locals() else 'unknown'}): {e}")
                # For UK securities, assume pence if we can't determine otherwise
                # This is a reasonable default for most UK stocks
                return True  # Conservative approach - convert if unsure
        
        def convert_price_if_needed(price, isin, ticker, trading_currency=None):
            """Convert price from pence to pounds if needed for UK securities."""
            if is_uk_security_in_pence(isin, ticker, trading_currency):
                return price / 100.0
            return price
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                
                # First pass: build ticker mappings
                for row in reader:
                    ticker = row.get("Ticker", "")
                    isin = row.get("ISIN", "")
                    currency = row.get("Currency (Price / share)", "")
                    if ticker and isin:
                        ticker_to_isin[ticker] = isin
                        if currency:
                            ticker_to_currency[ticker] = currency
                
                # Reset file pointer for second pass
                csvfile.seek(0)
                reader = csv.DictReader(csvfile)
                
                # Second pass: process transactions with price conversion
                for row in reader:
                    action = row.get("Action", "")
                    if action.lower() in ["market buy", "market sell", "stop buy", "stop sell", "limit buy", "limit sell"]:
                        try:
                            qty = float(row.get("No. of shares", "0") or 0)
                            price = float(row.get("Price / share", "0") or 0)
                            total = float(row.get("Total", "0") or 0)
                        except ValueError:
                            continue
                            
                        order_type = "Buy" if "buy" in action.lower() else "Sell"
                        ticker = row.get("Ticker", "")
                        
                        # Handle T212 ticker format - remove trailing 'l' if present
                        clean_ticker = ticker
                        if ticker.endswith('l') and len(ticker) > 1:
                            clean_ticker = ticker[:-1]
                        
                        # Get ISIN and currency for this ticker
                        isin = ticker_to_isin.get(clean_ticker, "")
                        trading_currency = ticker_to_currency.get(clean_ticker, "")
                        
                        # Convert price from pence to pounds if needed
                        converted_price = convert_price_if_needed(price, isin, clean_ticker, trading_currency)
                        
                        transactions_info.append({
                            "dateTime": row.get("Time", ""),
                            "ticker": clean_ticker,  # Use clean ticker for display
                            "name": row.get("Name", ""),
                            "orderType": order_type,
                            "quantity": qty,
                            "pricePerUnit": converted_price,
                            "totalValue": total,
                            "currency": row.get("Currency (Total)", ""),
                            "result": row.get("Result", "0")
                        })
        
        transactions_info.sort(key=lambda x: x.get("dateTime", ""), reverse=True)
        start_col, start_row = 2, 2
        
        # Add search instructions
        instruction_cell = self.ws.cell(row=start_row, column=start_col)
        instruction_cell.value = "💡 Use Excel's filter buttons in the header row to search and filter transactions"
        instruction_cell.font = Font(italic=True, size=12.5)
        instruction_cell.fill = PatternFill(start_color="fff6a8", end_color="fff6a8", fill_type="solid")
        self.ws.merge_cells(start_row=start_row, start_column=start_col, 
                            end_row=start_row, end_column=start_col + 6)
        
        # Title
        title_row = start_row + 1
        title_range = f"B{title_row}:H{title_row}"
        self.ws.merge_cells(title_range)
        title_cell = self.ws[f'B{title_row}']
        title_cell.value = "Detailed Transaction History"
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        # Headers
        headers = ["Date", "Ticker", "Asset Name", "Order Type", "Quantity", "Price/Unit", "Total Value"]
        header_row = title_row + 1
        for col_offset, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + col_offset, value=header)
            cell.fill = self.styles["grey"]
            cell.border = self.styles["table_border"]
            cell.font = Font(bold=True)
        
        # Data rows
        row = header_row + 1
        for tx in transactions_info:
            date = self.extract_date(tx.get("dateTime", ""))
            ticker = tx.get("ticker", "N/A")
            name = tx.get("name", "N/A")
            order_type = tx.get("orderType", "N/A")
            quantity = round(tx.get("quantity", 0), 4)
            price_per_unit = round(tx.get("pricePerUnit", 0), 4)
            total_value = round(tx.get("totalValue", 0), 2)
            
            row_fill = self.styles["green"] if order_type == "Buy" else self.styles["red"] if order_type == "Sell" else self.styles["grey"]
            values = [date, ticker, name, order_type, quantity, price_per_unit, total_value]
            
            for col_offset, val in enumerate(values):
                cell = self.ws.cell(row=row, column=start_col + col_offset, value=val)
                cell.border = self.styles["table_border"]
                if col_offset <= 2:  # Date, Ticker, Name columns
                    cell.fill = self.styles["grey"]
                elif col_offset == 3:  # Order Type column
                    cell.fill = row_fill
                    cell.font = Font(bold=True)
                else:  # Value columns
                    cell.fill = row_fill
            row += 1
        
        # Set column widths
        column_widths = {'B': 15, 'C': 12, 'D': 35, 'E': 12, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            self.ws.column_dimensions[col_letter].width = width
        
        last_data_row = row - 1 if transactions_info else header_row
        
        # Apply table border
        self.apply_table_border(ws=self.ws, first_row=start_row, last_row=last_data_row, 
                               first_col=start_col, last_col=start_col + len(headers) - 1)
        
        # Add Excel AutoFilter
        if transactions_info:
            filter_range = f"B{header_row}:E{last_data_row}"
            self.ws.auto_filter.ref = filter_range

    def wait_times_analysis(self):
        buy_transactions = {}
        sell_transactions = []
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row_data in reader:
                    action = row_data.get("Action", "")
                    if action.lower() in ["market buy", "market sell", "stop buy", "stop sell", "limit buy", "limit sell"]:
                        time_str = row_data.get("Time", "")
                        ticker = row_data.get("Ticker", "")
                        name = row_data.get("Name", "")
                        
                        try:
                            qty = float(row_data.get("No. of shares", "0") or 0)
                        except ValueError:
                            continue
                            
                        if qty <= 0:
                            continue
                        
                        transaction = {
                            "dateTime": time_str, "ticker": ticker, 
                            "name": name, "quantity": qty, "action": action
                        }
                        
                        if "buy" in action.lower():
                            if ticker not in buy_transactions:
                                buy_transactions[ticker] = []
                            buy_transactions[ticker].append(transaction)
                        elif "sell" in action.lower():
                            sell_transactions.append(transaction)
        
        all_individual_holds = []
        
        for sell_tx in sell_transactions:
            sell_ticker = sell_tx["ticker"]
            if sell_ticker not in buy_transactions:
                continue
                
            sell_name = sell_tx["name"]
            sell_date = sell_tx["dateTime"]
            remaining_sell_qty = sell_tx["quantity"]
            
            sorted_buys = sorted(buy_transactions[sell_ticker], key=lambda x: x["dateTime"])
            
            for buy_tx_original in sorted_buys:
                if remaining_sell_qty <= 0:
                    break
                
                buy_idx = buy_transactions[sell_ticker].index(buy_tx_original)
                buy_qty_available = buy_transactions[sell_ticker][buy_idx]["quantity"]
                
                if buy_qty_available <= 0:
                    continue

                matched_qty = min(remaining_sell_qty, buy_qty_available)
                
                try:
                    buy_dt = datetime.strptime(buy_tx_original["dateTime"].split(" ")[0], "%Y-%m-%d")
                    sell_dt = datetime.strptime(sell_date.split(" ")[0], "%Y-%m-%d")
                    
                    hold_days = (sell_dt - buy_dt).days
                    if hold_days >= 0:
                        all_individual_holds.append({
                            'ticker': sell_ticker, 'name': sell_name, 'days': hold_days
                        })
                        
                    buy_transactions[sell_ticker][buy_idx]["quantity"] -= matched_qty
                    remaining_sell_qty -= matched_qty
                except ValueError:
                    continue
        
        avg_hold_days = 0
        top_3_longest_holds = []
        
        if all_individual_holds:
            avg_hold_days = sum(item['days'] for item in all_individual_holds) / len(all_individual_holds)
            
            ticker_longest_holds = {}
            for hold in all_individual_holds:
                ticker, days = hold['ticker'], hold['days']
                if ticker not in ticker_longest_holds or days > ticker_longest_holds[ticker]['days']:
                    ticker_longest_holds[ticker] = hold
            
            top_3_longest_holds = sorted(ticker_longest_holds.values(), 
                                         key=lambda x: x['days'], reverse=True)[:3]
        
        start_col, start_row = 10, 2
        
        # Title
        title_range = f"J{start_row}:L{start_row}"
        self.ws.merge_cells(title_range)
        title_cell = self.ws[f'J{start_row}']
        title_cell.value = "Hold Time Statistics"
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = self.styles["dark_grey"]
        
        for r in range(start_row, start_row + 1):
            for c in range(start_col, start_col + 3):
                self.ws.cell(row=r, column=c).border = self.styles["title_border"]
        
        row = start_row + 1
        
        # Average Hold Time row
        blue_fill = PatternFill(start_color="e6f3ff", end_color="e6f3ff", fill_type="solid")
        
        cells = [
            (start_col, "Average Hold Time", self.styles["grey"]),
            (start_col + 1, round(avg_hold_days, 1), blue_fill),
            (start_col + 2, "days", blue_fill)
        ]
        
        for col_offset, value, fill in cells:
            cell = self.ws.cell(row=row, column=col_offset, value=value)
            cell.fill = fill
            cell.border = self.styles["table_border"]
        
        row += 1

        # Sub-header and data for longest held assets
        if top_3_longest_holds:
            # Sub-header
            sub_header_range = f"{chr(64 + start_col)}{row}:{chr(64 + start_col + 2)}{row}"
            self.ws.merge_cells(sub_header_range)
            self.ws.cell(row=row, column=start_col, value="Longest Held Unique Assets:").font = Font(bold=True)
            self.ws.cell(row=row, column=start_col).fill = self.styles["grey"]
            
            for c in range(start_col, start_col + 3):
                self.ws.cell(row=row, column=c).border = self.styles["table_border"]
            
            row += 1

            # Data rows for top 3 longest held assets
            for item in top_3_longest_holds:
                cells = [
                    (start_col, item['ticker'], self.styles["grey"]),
                    (start_col + 1, item['days'], blue_fill),
                    (start_col + 2, "days", blue_fill)
                ]
                
                for col_offset, value, fill in cells:
                    cell = self.ws.cell(row=row, column=col_offset, value=value)
                    cell.fill = fill
                    cell.border = self.styles["table_border"]
                row += 1
        
        # Set column widths
        for col_letter, width in {'J': 35, 'K': 15, 'L': 10}.items():
            self.ws.column_dimensions[col_letter].width = width
        
        # Apply border to entire table
        thin_side = Side(style='thin')
        last_row = row - 1
        
        for r in range(start_row, last_row + 1):
            for c in range(start_col, start_col + 3):
                cell = self.ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == start_col else border.left,
                    right=thin_side if c == start_col + 2 else border.right,
                    top=thin_side if r == start_row else border.top,
                    bottom=thin_side if r == last_row else border.bottom
                )
                cell.border = new_border
                
        # Store the last row number for use by fee_analysis
        self.last_wait_times_row = last_row

    def fee_analysis(self):
        fee_breakdown = {}
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        # Add gap of 2 rows from wait_times_analysis
        start_row = getattr(self, 'last_wait_times_row', 0) + 2
        start_col = 10
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                fee_types = [
                    ("Deposit Fee", "Deposit fee"),
                    ("Currency Conversion", "Currency conversion fee"),
                    ("Stamp Duty Tax", "Stamp duty reserve tax"),
                    ("Withholding Tax", "Withholding tax")
                ]
                
                for row in reader:
                    for fee_name, fee_column in fee_types:
                        fee_value = row.get(fee_column, "0")
                        if fee_value and fee_value != "0":
                            try:
                                amount = abs(float(fee_value))
                                if amount > 0:
                                    if fee_name not in fee_breakdown:
                                        fee_breakdown[fee_name] = 0
                                    fee_breakdown[fee_name] += amount
                            except ValueError:
                                continue
        
        total_fees = sum(fee_breakdown.values()) if fee_breakdown else 0
        
        # Create title
        title_range = f"J{start_row}:L{start_row}"
        self.ws.merge_cells(title_range)
        title_cell = self.ws[f'J{start_row}']
        title_cell.value = "Fee Breakdown"
        title_cell.font = Font(bold=True)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        # Create headers
        headers = ["Fee Type", "Amount", "Currency"]
        header_row = start_row + 1
        for col_offset, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + col_offset, value=header)
            cell.fill = self.styles["grey"]
            cell.border = self.styles["table_border"]
            cell.font = Font(bold=True)
        
        # Add fee data rows
        row = header_row + 1
        for fee_type, amount in sorted(fee_breakdown.items()):
            values = [fee_type, round(amount, 2), "EUR"]
            for col_offset, val in enumerate(values):
                cell = self.ws.cell(row=row, column=start_col + col_offset, value=val)
                cell.border = self.styles["table_border"]
                # Make fee type column grey, amount and currency columns red
                if col_offset == 0:  # Fee Type column
                    cell.fill = self.styles["grey"]
                else:  # Amount and Currency columns
                    cell.fill = self.styles["red"]
            row += 1
        
        # Add total row if fees exist
        if fee_breakdown:
            for col_offset, val in enumerate(["TOTAL FEES", round(total_fees, 2), "EUR"]):
                cell = self.ws.cell(row=row, column=start_col + col_offset, value=val)
                cell.border = self.styles["table_border"]
                cell.font = Font(bold=True)
                # Make total fee label grey, amount and currency columns red
                if col_offset == 0:  # Fee Type column
                    cell.fill = self.styles["grey"]
                else:  # Amount and Currency columns
                    cell.fill = self.styles["red"]
        
        # Set column widths
        for col_letter, width in {'J': 18, 'K': 12, 'L': 12}.items():
            self.ws.column_dimensions[col_letter].width = width
        
        # Apply border to entire table
        thin_side = Side(style='thin')
        last_data_row = row if fee_breakdown else header_row
        last_col = start_col + len(headers) - 1
        
        for r in range(start_row, last_data_row + 1):
            for c in range(start_col, last_col + 1):
                cell = self.ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == start_col else border.left,
                    right=thin_side if c == last_col else border.right,
                    top=thin_side if r == start_row else border.top,
                    bottom=thin_side if r == last_data_row else border.bottom
                )
                cell.border = new_border
        
        # Store the last row number for use by trading_statistics_analysis
        self.last_fee_row = last_data_row
                
    def capital_gains_graph(self):
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        capital_gains_data = defaultdict(float)
        
        if os.path.exists(csv_path):
            try:
                with open(csv_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    
                    for row in reader:
                        action = row.get("Action", "").strip()
                        time_str = row.get("Time", "").strip()
                        result_str = row.get("Result", "0").strip()
                        
                        if not time_str:
                            continue
                            
                        try:
                            date_part = time_str.split(" ")[0]
                            if not date_part:
                                continue
                                
                            date = datetime.strptime(date_part, "%Y-%m-%d")
                            
                            if action.lower() in ["market buy", "market sell", "stop buy", "stop sell", "limit buy", "limit sell"]:
                                if result_str and result_str not in ["0", ""]:
                                    try:
                                        result = float(result_str)
                                        if abs(result) > 0.01:
                                            capital_gains_data[date] += result
                                    except (ValueError, TypeError):
                                        continue
                                        
                        except (ValueError, IndexError):
                            continue
                            
            except Exception:
                pass
        
        # Create capital gains graph
        plt.style.use('default')
        fig, ax = plt.subplots(1, 1, figsize=(12, 6))
        fig.patch.set_facecolor('white')
        
        primary_color = '#4472C4'
        accent_color = '#FFC000'
        
        if capital_gains_data:
            dates = sorted(capital_gains_data.keys())
            cumulative_gains = []
            running_total = 0
            
            for date in dates:
                running_total += capital_gains_data[date]
                cumulative_gains.append(running_total)
            
            ax.plot(dates, cumulative_gains, color=primary_color, linewidth=3, 
                    marker='o', markersize=5, zorder=3)
            ax.fill_between(dates, cumulative_gains, alpha=0.3, color=primary_color, zorder=2)
            ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5, zorder=1)
            
            ax.set_title('Capital Gains Progress Over Time', fontsize=16, fontweight='bold', 
                         pad=20, color='#2F4F4F')
            ax.set_ylabel('Capital Gains (€)', fontsize=13, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
            
            if cumulative_gains:
                final_value = cumulative_gains[-1]
                max_value = max(cumulative_gains)
                
                ax.annotate(f'Current: €{final_value:.2f}', 
                           xy=(dates[-1], final_value), xytext=(20, 20), 
                           textcoords='offset points',
                           bbox=dict(boxstyle='round,pad=0.5', facecolor=primary_color, alpha=0.8),
                           fontsize=11, color='white', fontweight='bold',
                           arrowprops=dict(arrowstyle='->', color=primary_color, lw=2))
                
                if max_value != final_value:
                    max_idx = cumulative_gains.index(max_value)
                    ax.annotate(f'Peak: €{max_value:.2f}', 
                               xy=(dates[max_idx], max_value), xytext=(10, -30), 
                               textcoords='offset points',
                               bbox=dict(boxstyle='round,pad=0.3', facecolor=accent_color, alpha=0.7),
                               fontsize=9, color='black', fontweight='bold')
        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('gray')
        ax.spines['bottom'].set_color('gray')
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        ax.tick_params(colors='gray', which='both')
        ax.set_facecolor('#FAFAFA')
        
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', 
                   facecolor='white', edgecolor='none', pad_inches=0.2)
        img_buffer.seek(0)
        plt.close()
        
        img = Image(img_buffer)
        img.width = 720
        img.height = 288
        start_row = 2
        self.ws.add_image(img, f'N{start_row}')
    
    def dividends_graph(self):
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        dividend_data = defaultdict(float)
        
        if os.path.exists(csv_path):
            try:
                with open(csv_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    
                    for row in reader:
                        action = row.get("Action", "").strip()
                        time_str = row.get("Time", "").strip()
                        
                        if not time_str:
                            continue
                            
                        try:
                            date_part = time_str.split(" ")[0]
                            if not date_part:
                                continue
                                
                            date = datetime.strptime(date_part, "%Y-%m-%d")
                            
                            if "dividend" in action.lower():
                                total_str = row.get("Total", "0").strip()
                                try:
                                    dividend_amount = float(total_str)
                                    if dividend_amount > 0:
                                        dividend_data[date] += dividend_amount
                                except (ValueError, TypeError):
                                    continue
                                    
                        except (ValueError, IndexError):
                            continue
                            
            except Exception:
                pass
        
        # Create dividends graph
        plt.style.use('default')
        fig, ax = plt.subplots(1, 1, figsize=(12, 6))
        fig.patch.set_facecolor('white')
        
        secondary_color = '#70AD47'
        
        if dividend_data:
            dates = sorted(dividend_data.keys())
            cumulative_dividends = []
            running_total = 0
            
            for date in dates:
                running_total += dividend_data[date]
                cumulative_dividends.append(running_total)
            
            ax.plot(dates, cumulative_dividends, color=secondary_color, linewidth=3, 
                    marker='s', markersize=5, zorder=3)
            ax.fill_between(dates, cumulative_dividends, alpha=0.3, color=secondary_color, zorder=2)
            
            ax.set_title('Cumulative Dividend Growth Over Time', fontsize=16, fontweight='bold', 
                         pad=20, color='#2F4F4F')
            ax.set_ylabel('Total Dividends (€)', fontsize=13, fontweight='bold')
            ax.set_xlabel('Date', fontsize=13, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
            
            if cumulative_dividends:
                final_value = cumulative_dividends[-1]
                total_payments = len(dividend_data)
                
                ax.annotate(f'Total: €{final_value:.2f}', 
                           xy=(dates[-1], final_value), xytext=(20, 20), 
                           textcoords='offset points',
                           bbox=dict(boxstyle='round,pad=0.5', facecolor=secondary_color, alpha=0.8),
                           fontsize=11, color='white', fontweight='bold',
                           arrowprops=dict(arrowstyle='->', color=secondary_color, lw=2))
                
                ax.text(0.02, 0.98, f'Total: {total_payments} dividend payments', 
                        transform=ax.transAxes, ha='left', va='top',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8),
                        fontsize=10, fontweight='bold')
        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('gray')
        ax.spines['bottom'].set_color('gray')
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        ax.tick_params(colors='gray', which='both')
        ax.set_facecolor('#FAFAFA')
        
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', 
                   facecolor='white', edgecolor='none', pad_inches=0.2)
        img_buffer.seek(0)
        plt.close()
        
        img = Image(img_buffer)
        img.width = 720
        img.height = 288
        start_row = 20  # Position below capital gains graph
        self.ws.add_image(img, f'N{start_row}')
        
    def win_loss_statistics(self):
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        total_trades = winning_trades = 0
        total_pnl = 0.0
        start_row = getattr(self, 'last_fee_row', 0) + 2
        start_col = 10
        
        if os.path.exists(csv_path):
            try:
                with open(csv_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        action = row.get("Action", "").strip()
                        result_str = row.get("Result", "0").strip()
                        
                        if action.lower() in ["market buy", "market sell", "stop buy", "stop sell", "limit buy", "limit sell"]:
                            if result_str and result_str not in ["0", ""]:
                                try:
                                    result = float(result_str)
                                    if abs(result) > 0.01:
                                        total_trades += 1
                                        total_pnl += result
                                        if result > 0:
                                            winning_trades += 1
                                except (ValueError, TypeError):
                                    continue
            except Exception:
                pass
        
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        avg_pnl = total_pnl / total_trades if total_trades > 0 else 0
        
        title_range = f"J{start_row}:L{start_row}"
        self.ws.merge_cells(title_range)
        title_cell = self.ws[f'J{start_row}']
        title_cell.value = "Win/Loss Statistics"
        title_cell.font = Font(bold=True)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        headers = ["Metric", "Value", "Unit"]
        header_row = start_row + 1
        for col_offset, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + col_offset, value=header)
            cell.fill = self.styles["grey"]
            cell.border = self.styles["table_border"]
            cell.font = Font(bold=True)
        
        statistics_data = [
            ("Total Trades", total_trades, "trades"),
            ("Winning Trades", winning_trades, "trades"),
            ("Win Rate", round(win_rate, 2), "%"),
            ("Average P/L per Trade", round(avg_pnl, 2), "EUR")
        ]
        
        row = header_row + 1
        for metric, value, unit in statistics_data:
            for col_offset, val in enumerate([metric, value, unit]):
                cell = self.ws.cell(row=row, column=col_offset + start_col, value=val)
                cell.border = self.styles["table_border"]
                
                if col_offset == 0:
                    cell.fill = self.styles["grey"]
                elif col_offset in [1, 2]:
                    if metric == "Win Rate":
                        cell.fill = self.styles["green"] if value >= 50 else self.styles["red"]
                    elif metric == "Average P/L per Trade":
                        if value > 0:
                            cell.fill = self.styles["green"]
                        elif value < 0:
                            cell.fill = self.styles["red"]
                        else:
                            cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
            row += 1
        
        for col_letter, width in {'J': 18, 'K': 12, 'L': 12}.items():
            self.ws.column_dimensions[col_letter].width = width
        
        thin_side = Side(style='thin')
        last_data_row = row - 1
        last_col = start_col + len(headers) - 1
        
        for r in range(start_row, last_data_row + 1):
            for c in range(start_col, last_col + 1):
                cell = self.ws.cell(row=r, column=c)
                border = cell.border
                new_border = Border(
                    left=thin_side if c == start_col else border.left,
                    right=thin_side if c == last_col else border.right,
                    top=thin_side if r == start_row else border.top,
                    bottom=thin_side if r == last_data_row else border.bottom
                )
                cell.border = new_border

    def generate_sheet(self):
        self.order_history()
        self.wait_times_analysis()
        self.fee_analysis()
        self.win_loss_statistics()
        self.capital_gains_graph()
        self.dividends_graph()
