import os
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from datetime import datetime
from AccountSummary import AccountSummary
from AdvancedAccountInfo import AdvancedAccountInfo

CACHE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache")

def load_cached(name, fallback_func):
    path = os.path.join(CACHE_DIR, f"{name}.json")
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return fallback_func()

# Helper functions for Excel operations

def create_title(ws, title_text, title_range, fill_color=None, font_bold=True, font_size=None):
    """Create a styled title for a table section"""
    ws.merge_cells(title_range)
    
    # Extract first cell reference from the range
    first_cell_ref = title_range.split(':')[0]
    title_cell = ws[first_cell_ref]
    
    title_cell.value = title_text
    if font_bold:
        title_cell.font = Font(bold=True, size=font_size) if font_size else Font(bold=True)
    if fill_color:
        title_cell.fill = fill_color
        
    # Apply border to all cells in the range
    for row in ws[title_range]:
        for cell in row:
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
    
    return title_cell

def create_headers(ws, headers, start_row, start_col, fill_color=None, border=None, font_bold=True):
    """Create styled headers for a table"""
    for col_offset, header in enumerate(headers):
        col = start_col + col_offset
        cell = ws.cell(row=start_row, column=col, value=header)
        if fill_color:
            cell.fill = fill_color
        if border:
            cell.border = border
        if font_bold:
            cell.font = Font(bold=True)
    
    return start_row + 1  # Return the next row index

def set_column_widths(ws, width_map):
    """Set column widths based on a dictionary mapping"""
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width

def apply_table_border(ws, first_row, last_row, first_col, last_col):
    """Apply full border to a table range"""
    thin_side = Side(style='thin')
    
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            border = cell.border
            new_border = Border(
                left=thin_side if c == first_col else border.left,
                right=thin_side if c == last_col else border.right,
                top=thin_side if r == first_row else border.top,
                bottom=thin_side if r == last_row else border.bottom,
            )
            cell.border = new_border

def extract_date(date_time_str):
    """Extract date part from a datetime string"""
    if " " in date_time_str:
        return date_time_str.split(" ")[0]
    elif "T" in date_time_str:
        return date_time_str.split("T")[0]
    else:
        return date_time_str

def read_csv_data(csv_path):
    """Read data from a CSV file safely"""
    data = []
    # Make path absolute if it's not already
    if not os.path.isabs(csv_path):
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), csv_path)
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                data.append(row)
    return data

def make_xslx():
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Summary"
    
    wb.create_sheet("Advanced Account Info")

    # Styles
    styles = {
        "table_border": Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        "title_border": Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        ),
        "dark_grey": PatternFill(start_color="9C9C9C", end_color="9C9C9C", fill_type="solid"),
        "grey": PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid"),
        "red": PatternFill(start_color="e8baba", end_color="e8baba", fill_type="solid"),
        "green": PatternFill(start_color="c3e8cb", end_color="c3e8cb", fill_type="solid")
    }

    # Create instances of both classes
    account_summary = AccountSummary(
        wb=wb, 
        ws=ws, 
        styles=styles, 
        load_cached_func=load_cached, 
        extract_date_func=extract_date, 
        apply_border_func=apply_table_border
    )
    
    advanced_account_info = AdvancedAccountInfo(
        wb=wb, 
        ws=wb["Advanced Account Info"], 
        styles=styles, 
        extract_date_func=extract_date, 
        apply_border_func=apply_table_border
    )
    
    # Generate both sheets
    account_summary.generate_sheet()
    advanced_account_info.generate_sheet()

    # Save the workbook
    wb.save("AccountAnalysis.xlsx")

if __name__ == "__main__":
    make_xslx()