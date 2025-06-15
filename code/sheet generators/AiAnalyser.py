import os
import csv
import json
import openai
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from datetime import datetime
import sys
from PIL import Image as PILImage, ImageDraw, ImageFont
import textwrap
import io
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from AccountData import get_open_positions, get_cash_info
from dotenv import load_dotenv

load_dotenv()

class AiAnalyser:
    def __init__(self, wb, styles, load_cached_func, apply_border_func):
        self.wb = wb
        self.ws = wb.create_sheet("AI Analysis")
        self.styles = styles
        self.load_cached = load_cached_func
        self.apply_table_border = apply_border_func
        self.client = None
        self.api_available = False
        
        self.test_api_connection()
    
    def test_api_connection(self):
        try:
            api_key = os.getenv("OPENAI_API_KEY")
            if not api_key:
                return False
            
            self.client = openai.OpenAI(api_key=api_key)
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": "Hello"}],
                max_tokens=5,
                temperature=0
            )
            
            self.api_available = True
            return True
            
        except Exception as e:
            self.api_available = False
            return False
    
    def load_raw_data(self):
        positions = self.load_cached("open_positions", get_open_positions)
        cash_info = self.load_cached("cash_info", get_cash_info)
        pies_info = self.load_cached("pies_info", lambda: {})
        
        trading_history = []
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                trading_history = list(reader)
        
        return {
            "positions": positions,
            "cash_info": cash_info,
            "pies_info": pies_info,
            "trading_history": trading_history
        }
    
    def _prepare_comprehensive_data(self, raw_data):
        data_dump = "=== RAW TRADING212 DATA ===\n\n"
        
        data_dump += "=== CASH_INFO.JSON ===\n"
        data_dump += json.dumps(raw_data["cash_info"], indent=2) + "\n\n"
        
        data_dump += "=== OPEN_POSITIONS.JSON ===\n"
        data_dump += json.dumps(raw_data["positions"], indent=2) + "\n\n"
        
        if raw_data["pies_info"]:
            data_dump += "=== PIES_INFO.JSON ===\n"
            data_dump += json.dumps(raw_data["pies_info"], indent=2) + "\n\n"
        
        data_dump += "=== TRADING212_HISTORY.CSV (Recent 50 entries) ===\n"
        data_dump += "Action,Time,ISIN,Ticker,Name,No. of shares,Price / share,Currency (Price / share),Exchange rate,Result,Currency (Result),Total,Currency (Total),Withholding tax,Currency (Withholding tax),Charge amount (per transaction),Currency (Charge),Finra fee,Currency (Finra fee),Stamp duty reserve tax,Currency (Stamp duty reserve tax),Notes,ID,Currency conversion fee,Currency (Currency conversion fee)\n"
        
        for i, trade in enumerate(raw_data["trading_history"][:50]):
            row_data = [
                trade.get("Action", ""),
                trade.get("Time", ""),
                trade.get("ISIN", ""),
                trade.get("Ticker", ""),
                trade.get("Name", ""),
                trade.get("No. of shares", ""),
                trade.get("Price / share", ""),
                trade.get("Currency (Price / share)", ""),
                trade.get("Exchange rate", ""),
                trade.get("Result", ""),
                trade.get("Currency (Result)", ""),
                trade.get("Total", ""),
                trade.get("Currency (Total)", ""),
                trade.get("Withholding tax", ""),
                trade.get("Currency (Withholding tax)", ""),
                trade.get("Charge amount (per transaction)", ""),
                trade.get("Currency (Charge)", ""),
                trade.get("Finra fee", ""),
                trade.get("Currency (Finra fee)", ""),
                trade.get("Stamp duty reserve tax", ""),
                trade.get("Currency (Stamp duty reserve tax)", ""),
                trade.get("Notes", ""),
                trade.get("ID", ""),
                trade.get("Currency conversion fee", ""),
                trade.get("Currency (Currency conversion fee)", "")
            ]
            data_dump += ",".join([str(item) for item in row_data]) + "\n"
        
        data_dump += f"\n=== TOTAL TRADING HISTORY ENTRIES: {len(raw_data['trading_history'])} ===\n"
        
        return data_dump

    def get_ai_insights(self, raw_data):
        if not self.api_available:
            return {"AI Portfolio Analysis": "AI analysis unavailable - OpenAI API connection failed"}
        
        try:
            comprehensive_data = self._prepare_comprehensive_data(raw_data)
            
            prompt = f"""
            You are a professional portfolio analyst with 20+ years of experience. Analyze this complete Trading212 portfolio data and provide comprehensive, detailed feedback.

            COMPLETE PORTFOLIO DATA:
            {comprehensive_data}

            Please provide a DETAILED analysis covering these areas with specific insights and actionable recommendations:

            1. CASH ALLOCATION & LIQUIDITY MANAGEMENT:
            - Analyze the current cash position and liquidity levels
            - Evaluate cash-to-investment ratio efficiency
            - Identify opportunities for better cash deployment
            - Assess emergency fund adequacy

            2. PORTFOLIO CONCENTRATION & DIVERSIFICATION RISKS:
            - Calculate and evaluate position concentration levels
            - Identify potential single-stock risk exposures
            - Analyze sector/geographic diversification
            - Recommend specific diversification improvements

            3. TRADING PATTERNS & PERFORMANCE METRICS:
            - Evaluate trading frequency and timing patterns
            - Analyze win/loss ratios and trade profitability
            - Assess fee impact on overall returns
            - Identify behavioral trading patterns (good and problematic)

            4. PORTFOLIO HEALTH & STRATEGIC RECOMMENDATIONS:
            - Overall portfolio performance assessment
            - Risk-adjusted return analysis
            - Strategic asset allocation recommendations
            - Specific action items for portfolio improvement

            5. DETAILED OBSERVATIONS:
            - Notable holdings analysis
            - Market timing insights from trading history
            - Potential tax implications
            - Long-term wealth building recommendations

            Please be SPECIFIC with numbers, percentages, and concrete recommendations. Reference actual positions, amounts, and dates from the data. Provide actionable insights that the investor can implement immediately.

            Make your analysis comprehensive and detailed - aim for 600-800 words covering all aspects thoroughly.
            """
            
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1200,
                temperature=0.3
            )
            
            return {"AI Portfolio Analysis": response.choices[0].message.content.strip()}
            
        except Exception as e:
            return {"AI Portfolio Analysis": "AI analysis unavailable - Error occurred during analysis"}
    
    def create_text_image(self, text, width=1000, font_size=14):
        """Create a high-quality image from text with proper word wrapping"""
        # Set up basic parameters for better quality
        line_height = int(font_size * 1.5)  # Increased line spacing
        padding = 30  # Increased padding
        background_color = (255, 255, 255)  # Pure white background for better readability
        text_color = (33, 37, 41)  # Dark text
        
        # Try to use a system font, fallback to default if not available
        try:
            # Try common system fonts with larger size
            font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", font_size)
        except (OSError, IOError):
            try:
                font = ImageFont.truetype("/System/Library/Fonts/Arial.ttf", font_size)
            except (OSError, IOError):
                try:
                    font = ImageFont.truetype("/Library/Fonts/Arial.ttf", font_size)
                except (OSError, IOError):
                    font = ImageFont.load_default()
        
        # Wrap text to fit the width with better character estimation
        chars_per_line = int(width / (font_size * 0.55))  # Better estimation for readability
        wrapper = textwrap.TextWrapper(width=chars_per_line, break_long_words=False, break_on_hyphens=False)
        wrapped_lines = []
        
        # Process paragraphs separately for better formatting
        paragraphs = text.split('\n\n')
        for i, paragraph in enumerate(paragraphs):
            if paragraph.strip():
                # Handle existing line breaks within paragraphs
                para_lines = paragraph.split('\n')
                for line in para_lines:
                    if line.strip():
                        wrapped = wrapper.wrap(line.strip())
                        wrapped_lines.extend(wrapped)
                
                # Add spacing between paragraphs (except for the last one)
                if i < len(paragraphs) - 1:
                    wrapped_lines.append('')
        
        # Calculate image height with extra space
        height = (len(wrapped_lines) * line_height) + (2 * padding) + 50  # Extra 50px buffer
        
        # Create high-resolution image
        img = PILImage.new('RGB', (width, height), background_color)
        draw = ImageDraw.Draw(img)
        
        # Draw text with anti-aliasing
        y_position = padding
        for line in wrapped_lines:
            if line:  # Only draw non-empty lines
                draw.text((padding, y_position), line, font=font, fill=text_color)
            y_position += line_height
        
        # Save to bytes buffer with high quality
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG', optimize=False, quality=100)
        img_buffer.seek(0)
        
        return img_buffer
        
    def create_insights_table(self, insights, start_row=2):
        start_col = 2
        
        # Create title row
        title_range = f"B{start_row}:I{start_row}"
        self.ws.merge_cells(title_range)
        title_cell = self.ws.cell(row=start_row, column=start_col, value="AI Portfolio Analysis & Recommendations")
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = self.styles["dark_grey"]
        title_cell.alignment = Alignment(horizontal="center")
        
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        # Create image from AI analysis text
        analysis_content = insights.get("AI Portfolio Analysis", "No analysis available")
        
        # Generate image from text
        img_buffer = self.create_text_image(analysis_content, width=900, font_size=13)
        
        # Insert image into worksheet
        img = Image(img_buffer)
        
        # Position image starting from row after title
        img.anchor = f"B{start_row + 2}"
        
        # Add image to worksheet
        self.ws.add_image(img)
        
        # Set column widths to accommodate the image
        column_widths = {
            'B': 12, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 
            'G': 15, 'H': 15, 'I': 15
        }
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Return a reasonable next row (estimate based on image height)
        # Assuming roughly 600px image height = ~30 rows at 20px per row
        estimated_image_rows = 30
        return start_row + 2 + estimated_image_rows
    
    def generate_sheet(self):
        raw_data = self.load_raw_data()
        insights = self.get_ai_insights(raw_data)
        self.create_insights_table(insights, start_row=2)