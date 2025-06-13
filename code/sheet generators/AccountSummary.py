import os
import csv
import os
import csv
from openpyxl.styles import Font, PatternFill, Border, Side
from AccountData import get_cash_info, get_open_positions, get_pies

class AccountSummary:
    def __init__(self, wb, ws, styles, load_cached_func, extract_date_func, apply_border_func):
        self.wb = wb
        self.ws = ws
        self.styles = styles
        self.load_cached = load_cached_func
        self.extract_date = extract_date_func
        self.apply_table_border = apply_border_func
        
    def cash_info_table(self):
        cash_info = self.load_cached("cash_info", get_cash_info)
        title_range = "B2:D2"
        self.ws.merge_cells(title_range)
        title_cell = self.ws['B2']
        title_cell.value = "Cash Info"
        title_cell.font = Font(bold=True)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]

        label_map = {
            "total": "Total Account Value",
            "free": "Cash",
            "invested": "Currently Invested",
            "blocked": "Blocked Amount",
            "pieCash": "Cash in Pies",
            "result": "Realised Return",
            "ppl": "Open Position P/L",
        }

        row = 3
        for key, label in label_map.items():
            self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            label_cell = self.ws.cell(row=row, column=2, value=label)
            value = cash_info.get(key) or 0
            value_cell = self.ws.cell(row=row, column=4, value=value)
            
            cell_fill = self.styles["grey"]
            if key in ["result", "ppl"] and isinstance(value, (int, float)):
                if value > 0:
                    cell_fill = self.styles["green"]
                elif value < 0:
                    cell_fill = self.styles["red"]
                    
            label_cell.fill = cell_fill
            value_cell.fill = cell_fill
            
            for col in range(2, 5):
                self.ws.cell(row=row, column=col).border = self.styles["table_border"]
            row += 1
        
        self.apply_table_border(self.ws, 2, row - 1, 2, 4)

    def open_positions_table(self):
        positions = self.load_cached("open_positions", get_open_positions)
        positions = sorted(positions, key=lambda x: x.get("ppl", 0), reverse=True)
        start_col, start_row = 6, 2
        
        # Title
        title_range = "F2:K2"
        self.ws.merge_cells(title_range)
        title_cell = self.ws['F2']
        title_cell.value = "Open Positions"
        title_cell.font = Font(bold=True)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        # Headers
        headers = ["Ticker", "Quantity", "Avg. Price", "Current Price", "P/L", "FX P/L"]
        header_row = start_row + 1
        for col_offset, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + col_offset, value=header)
            cell.fill = self.styles["grey"]
            cell.border = self.styles["table_border"]
            cell.font = Font(bold=True)
        
        # Data rows
        row = header_row + 1
        for pos in positions:
            ticker = pos.get("ticker", "N/A").split("_")[0]
            quantity = round(pos.get("quantity") or 0.0, 2)
            avg_price = round(pos.get("averagePrice") or 0.0, 2)
            current_price = round(pos.get("currentPrice") or 0.0, 2)
            ppl = round(pos.get("ppl") or 0.0, 2)
            fx_ppl = round(pos.get("fxPpl") or 0.0, 2)
            values = [ticker, quantity, avg_price, current_price, ppl, fx_ppl]
            row_fill = self.styles["green"] if ppl > 0 else self.styles["red"] if ppl < 0 else self.styles["grey"]
            
            for col_offset, val in enumerate(values):
                cell = self.ws.cell(row=row, column=start_col + col_offset, value=val)
                cell.border = self.styles["table_border"]
                
                if col_offset == 0:
                    cell.fill = self.styles["grey"]
                elif col_offset == 5:  # FX P/L column
                    cell.fill = self.styles["green"] if fx_ppl > 0 else self.styles["red"] if fx_ppl < 0 else self.styles["grey"]
                else:
                    cell.fill = row_fill
            row += 1
        
        # Set column widths and apply border
        for col_letter in ['F', 'G', 'H', 'I', 'J', 'K']:
            self.ws.column_dimensions[col_letter].width = 15
        
        self.apply_table_border(self.ws, start_row, row - 1, start_col, start_col + len(headers) - 1)

    def historical_transactions(self):
        transactions_info = []
        csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "cache", "trading212_history.csv")
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    action = row.get("Action", "")
                    if action.lower() in ["deposit", "withdrawal", "withdraw"]:
                        time_str = row.get("Time", "")
                        total = row.get("Total", "0")
                        
                        try:
                            amount = float(total) if total else 0
                        except ValueError:
                            amount = 0
                        
                        transactions_info.append({
                            "dateTime": time_str,
                            "type": action,
                            "amount": amount
                        })
        
        start_col, start_row = 2, 11
        
        # Title
        title_range = "B11:D11"
        self.ws.merge_cells(title_range)
        title_cell = self.ws['B11']
        title_cell.value = "Total Transactions"
        title_cell.font = Font(bold=True)
        title_cell.fill = self.styles["dark_grey"]
        for row in self.ws[title_range]:
            for cell in row:
                cell.border = self.styles["title_border"]
        
        # Headers
        headers = ["Date", "Transaction Type", "Amount"]
        header_row = start_row + 1
        for col_offset, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + col_offset, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.styles["grey"]
            cell.border = self.styles["table_border"]
        
        # Data rows
        row = header_row + 1
        for tx in transactions_info:
            tx_type = tx.get("type", "N/A")
            if str(tx_type).lower() == "fee":
                continue
                
            date = self.extract_date(tx.get("dateTime", ""))
            amount = tx.get("amount", 0)
            tx_type_lower = str(tx_type).lower()
            
            row_fill = self.styles["grey"]
            value_fill = self.styles["green"] if tx_type_lower == "deposit" else self.styles["red"] if tx_type_lower in ["withdraw", "withdrawal"] else self.styles["grey"]
            
            values = [date, tx_type, amount]
            for col_offset, val in enumerate(values):
                cell = self.ws.cell(row=row, column=start_col + col_offset, value=val)
                cell.border = self.styles["table_border"]
                cell.fill = row_fill if col_offset == 0 else value_fill
            row += 1
        
        # Set column widths
        for col_letter in ['B', 'C', 'D']:
            self.ws.column_dimensions[col_letter].width = 15
        
        self.apply_table_border(self.ws, start_row, row - 1, start_col, start_col + len(headers) - 1)

    def pies_tables(self):
        pies_info = self.load_cached("pies_info", lambda: get_pies(include_detailed=True))
        start_col = 13  # Column M (beside open positions)
        start_row = 2
        
        for pie in pies_info:
            pie_id = pie.get("id", "")
            name = pie.get("detailed", {}).get("settings", {}).get("name") or pie.get("name", "N/A")
            detailed = pie.get("detailed", {})
            instruments = detailed.get("instruments", [])
            
            # Skip pies with no assets or value
            if not instruments:
                continue
            total_value = sum(inst.get("result", {}).get("priceAvgValue", 0) for inst in instruments)
            if total_value == 0:
                continue
            
            # Sort instruments by performance (greatest gains to greatest losses)
            instruments = sorted(instruments, key=lambda x: x.get("result", {}).get("priceAvgResultCoef", 0), reverse=True)
            
            # Get pie summary data
            pie_result = pie.get("result", {})
            total_invested = round(pie_result.get("priceAvgInvestedValue", 0), 2)
            pie_pl = round(pie_result.get("priceAvgResult", 0), 2)
            pie_pl_percent = round(pie_result.get("priceAvgResultCoef", 0) * 100, 2)
            
            # Title
            title_range = self.ws.cell(row=start_row, column=start_col).coordinate + ":" + self.ws.cell(row=start_row, column=start_col+4).coordinate
            self.ws.merge_cells(title_range)
            title_cell = self.ws.cell(row=start_row, column=start_col)
            title_cell.value = f"Pie: {name} (ID: {pie_id})"
            title_cell.font = Font(bold=True)
            title_cell.fill = self.styles["dark_grey"]
            for col in range(start_col, start_col+5):
                cell = self.ws.cell(row=start_row, column=col)
                cell.border = self.styles["title_border"]
            
            # Holdings header
            holdings_row = start_row + 1
            subheaders = ["Ticker", "Weight %", "Performance %", "Quantity", "Value"]
            for col_offset, subheader in enumerate(subheaders):
                col = start_col + col_offset
                cell = self.ws.cell(row=holdings_row, column=col, value=subheader)
                cell.font = Font(bold=True)
                cell.fill = self.styles["grey"]
                cell.border = self.styles["table_border"]
            
            # Holdings data
            row = holdings_row + 1
            for inst in instruments:
                ticker = inst.get("ticker", "").split("_")[0]  # Clean ticker
                weight = round(inst.get("currentShare", 0) * 100, 2)
                perf = round(inst.get("result", {}).get("priceAvgResultCoef", 0) * 100, 2)
                qty = round(inst.get("ownedQuantity", 0), 4)
                value = round(inst.get("result", {}).get("priceAvgValue", 0), 2)
                values = [ticker, weight, perf, qty, value]
                row_fill = self.styles["green"] if perf > 0 else self.styles["red"] if perf < 0 else self.styles["grey"]
                
                for col_offset, val in enumerate(values):
                    col = start_col + col_offset
                    cell = self.ws.cell(row=row, column=col, value=val)
                    cell.border = self.styles["table_border"]
                    if col_offset == 0:  # Ticker column
                        cell.fill = self.styles["grey"]
                    else:
                        cell.fill = row_fill
                row += 1
            
            # Total value row
            total_row = row
            self.ws.cell(row=total_row, column=start_col, value="Total Pie Value:").font = Font(bold=True)
            self.ws.cell(row=total_row, column=start_col+4, value=round(total_value, 2)).font = Font(bold=True)
            for col in range(start_col, start_col+5):
                cell = self.ws.cell(row=total_row, column=col)
                cell.border = self.styles["table_border"]
                cell.fill = self.styles["grey"]
            
            # Pie Summary Section
            summary_row = total_row + 1
            summary_labels = ["Initial Investment:", "Pie P/L:", "P/L %:"]
            summary_values = [total_invested, pie_pl, pie_pl_percent]
            
            for i, (label, value) in enumerate(zip(summary_labels, summary_values)):
                # Create label in first column and value in last column
                self.ws.cell(row=summary_row + i, column=start_col, value=label).font = Font(bold=True)
                self.ws.cell(row=summary_row + i, column=start_col + 4, value=value).font = Font(bold=True)
                
                # Determine value fill color for P/L rows
                if i > 0:  # P/L and P/L % rows
                    if value > 0:
                        value_fill = self.styles["green"]
                    elif value < 0:
                        value_fill = self.styles["red"]
                    else:
                        value_fill = self.styles["grey"]
                else:
                    value_fill = self.styles["grey"]
                
                # Apply styling and borders to all cells in the row
                for col in range(start_col, start_col+5):
                    cell = self.ws.cell(row=summary_row + i, column=col)
                    cell.border = self.styles["table_border"]
                    if col == start_col + 4:  # Value column
                        cell.fill = value_fill
                    else:  # Label and middle columns
                        cell.fill = self.styles["grey"]
            
            last_summary_row = summary_row + len(summary_labels) - 1
            
            # Apply full table border
            thin_side = Side(style='thin')
            first_row = start_row
            last_data_row = last_summary_row
            first_col = start_col
            last_col = start_col + 4
            for r in range(first_row, last_data_row + 1):
                for c in range(first_col, last_col + 1):
                    cell = self.ws.cell(row=r, column=c)
                    border = cell.border
                    new_border = Border(
                        left=thin_side if c == first_col else border.left,
                        right=thin_side if c == last_col else border.right,
                        top=thin_side if r == first_row else border.top,
                        bottom=thin_side if r == last_data_row else border.bottom,
                    )
                    cell.border = new_border
            
            # Set column widths
            for col_idx in range(start_col, start_col+5):
                col_letter = chr(64 + col_idx)
                self.ws.column_dimensions[col_letter].width = 15
            
            start_row = last_summary_row + 3  # Add space between pies
            
    def generate_sheet(self):
        self.cash_info_table()
        self.open_positions_table()
        self.historical_transactions()
        self.pies_tables()
